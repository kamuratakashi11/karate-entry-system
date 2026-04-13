import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import json
import datetime
import io
import os
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import time
import random
import base64
import requests

# 安全なインポート
try:
    from openpyxl.cell import MergedCell
except ImportError:
    try:
        from openpyxl.cell.cell import MergedCell
    except ImportError:
        class MergedCell: pass

# ---------------------------------------------------------
# 1. 定数・初期設定・ヘルパー
# ---------------------------------------------------------
KEY_FILE = 'secrets.json'
SHEET_NAME = 'tournament_db' 
V2_PREFIX = "v2_" 

GAS_WEBAPP_URL = "https://script.google.com/macros/s/AKfycbwTegYveIaIKagvcsJBcLlxbjVx7siHoeUmh_3YrRSu9uOpvl6Uo8X3NifGinnzuxSA/exec"

MEMBERS_COLS = ["school_id", "name", "sex", "grade", "dob", "jkf_no", "display_order", "active"]

def to_half_width(text):
    if not text: return ""
    return str(text).translate(str.maketrans('０１２３４５６７８９', '0123456789')).strip()

def to_safe_int(val):
    try:
        s = to_half_width(str(val))
        return int(s)
    except: return 999

def generate_school_id():
    return f"sch_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}"

DEFAULT_TOURNAMENTS = {
    "kantou": {"name": "関東高等学校空手道大会 埼玉県予選", "template": "template_kantou.xlsx", "type": "standard", "grades": [1, 2, 3], "active": True},
    "interhigh": {"name": "インターハイ 埼玉県予選", "template": "template_interhigh.xlsx", "type": "standard", "grades": [1, 2, 3], "active": False},
    "shinjin": {"name": "新人大会", "template": "template_shinjin.xlsx", "type": "shinjin", "grades": [1, 2], "weights_m": "-55,-61,-68,-76,+76", "weights_w": "-48,-53,-59,-66,+66", "active": False},
    "senbatsu": {"name": "全国選抜 埼玉県予選", "template": "template_senbatsu.xlsx", "type": "division", "grades": [1, 2], "active": False}
}

DEFAULT_LIMITS = {
    "team_kata": {"min": 3, "max": 3},
    "team_kumite_5": {"min": 3, "max": 5},
    "team_kumite_3": {"min": 2, "max": 3},
    "ind_kata_reg": {"max": 4}, "ind_kata_sub": {"max": 2},
    "ind_kumi_reg": {"max": 4}, "ind_kumi_sub": {"max": 2}
}

COORD_DEF = {
    "year": "E3", "tournament_name": "I3", "date": "M7",
    "school_name": "C8", "principal": "C9", "head_advisor": "O9",
    "advisors": [{"name": "B42", "d1": "C42", "d2": "F42"}, {"name": "B43", "d1": "C43", "d2": "F43"}, {"name": "K42", "d1": "Q42", "d2": "U42"}, {"name": "K43", "d1": "Q43", "d2": "U43"}],
    "start_row": 16, "cap": 22, "offset": 46,
    "cols": {"name": 2, "grade": 3, "dob": 4, "jkf_no": 19, "m_team_kata": 11, "m_team_kumite": 12, "m_kata": 13, "m_kumite": 14, "w_team_kata": 15, "w_team_kumite": 16, "w_kata": 17, "w_kumite": 18}
}

# ---------------------------------------------------------
# 2. Google Sheets 接続
# ---------------------------------------------------------
@st.cache_resource
def get_gsheet_client():
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    if os.path.exists(KEY_FILE): creds = ServiceAccountCredentials.from_json_keyfile_name(KEY_FILE, scope)
    else:
        try:
            vals = st.secrets["gcp_key"]
            key_dict = json.loads(vals) if isinstance(vals, str) else vals
            creds = ServiceAccountCredentials.from_json_keyfile_dict(key_dict, scope)
        except Exception as e:
            st.error(f"認証設定エラー: {e}"); st.stop()
    return gspread.authorize(creds)

def retry_api(func):
    def wrapper(*args, **kwargs):
        for i in range(3):
            try: return func(*args, **kwargs)
            except Exception as e:
                if i == 2: raise e
                time.sleep(1 + random.random())
    return wrapper

@retry_api
def get_worksheet_safe(tab_name):
    client = get_gsheet_client()
    try: sh = client.open(SHEET_NAME)
    except: st.error(f"スプレッドシート '{SHEET_NAME}' が見つかりません。"); st.stop()
    try: ws = sh.worksheet(tab_name)
    except: 
        try: ws = sh.add_worksheet(title=tab_name, rows=100, cols=20)
        except: ws = sh.worksheet(tab_name)
    return ws

# ---------------------------------------------------------
# 3. データ操作 
# ---------------------------------------------------------
def load_json(tab_name, default):
    target_tab = f"{V2_PREFIX}{tab_name}"
    try:
        ws = get_worksheet_safe(target_tab); recs = ws.get_all_values()
        if not recs: return default
        if len(recs) == 1 and len(recs[0]) >= 1:
            val = str(recs[0][0])
            if val.startswith("{") or val.startswith("["): return json.loads(val) if json.loads(val) is not None else default
        result = {}
        for row in recs:
            if len(row) >= 2:
                key = row[0]; val_str = row[1]
                try: result[key] = json.loads(val_str)
                except: result[key] = val_str
        return result if result else default
    except: return default

def save_json(tab_name, data):
    target_tab = f"{V2_PREFIX}{tab_name}"; ws = get_worksheet_safe(target_tab)
    if not isinstance(data, dict):
        ws.clear(); ws.update_acell('A1', json.dumps(data, ensure_ascii=False)); return
    rows = [[str(k), json.dumps(v, ensure_ascii=False)] for k, v in data.items()]
    ws.clear()
    if rows: ws.update(rows)

def load_members_master(force_reload=False):
    if not force_reload and "v2_master_cache" in st.session_state: return st.session_state["v2_master_cache"]
    try:
        ws = get_worksheet_safe(f"{V2_PREFIX}members"); recs = ws.get_all_records()
        if not recs: df = pd.DataFrame(columns=MEMBERS_COLS)
        else:
             df = pd.DataFrame(recs)
             for c in MEMBERS_COLS:
                 if c not in df.columns: df[c] = ""
    except: return pd.DataFrame(columns=MEMBERS_COLS)
    df['grade'] = pd.to_numeric(df['grade'], errors='coerce').fillna(0).astype(int); df['jkf_no'] = df['jkf_no'].astype(str).replace('nan', ''); df['dob'] = df['dob'].astype(str).replace('nan', ''); df['display_order'] = df['display_order'].astype(str).replace('nan', '') 
    df = df[MEMBERS_COLS]; st.session_state["v2_master_cache"] = df; return df

def save_members_master(df):
    ws = get_worksheet_safe(f"{V2_PREFIX}members"); ws.clear()
    df = df.fillna(""); df['jkf_no'] = df['jkf_no'].astype(str); df['dob'] = df['dob'].astype(str); df['display_order'] = df['display_order'].astype(str) 
    for c in MEMBERS_COLS:
        if c not in df.columns: df[c] = ""
    df_to_save = df[MEMBERS_COLS]
    ws.update([df_to_save.columns.tolist()] + df_to_save.astype(str).values.tolist())
    st.session_state["v2_master_cache"] = df_to_save

def archive_graduates(grad_df, auth_data):
    if grad_df.empty: return
    ws_grad = get_worksheet_safe(f"{V2_PREFIX}graduates")
    grad_df = grad_df.copy(); grad_df["archived_school_name"] = grad_df["school_id"].apply(lambda sid: auth_data.get(sid, {}).get("base_name", "不明")); grad_df["archived_date"] = datetime.date.today().strftime("%Y-%m-%d")
    if not ws_grad.get_all_values(): ws_grad.append_row(grad_df.columns.tolist())
    ws_grad.append_rows(grad_df.astype(str).values.tolist())

def clear_graduates_archive(): get_worksheet_safe(f"{V2_PREFIX}graduates").clear()

def get_graduates_df():
    try:
        recs = get_worksheet_safe(f"{V2_PREFIX}graduates").get_all_records()
        return pd.DataFrame(recs) if recs else pd.DataFrame()
    except: return pd.DataFrame()

def load_entries(tournament_id, force_reload=False):
    key = f"v2_entry_cache_{tournament_id}"
    if not force_reload and key in st.session_state: return st.session_state[key]
    data = load_json(f"entry_{tournament_id}", {}); st.session_state[key] = data; return data

def save_entries(tournament_id, data):
    save_json(f"entry_{tournament_id}", data); st.session_state[f"v2_entry_cache_{tournament_id}"] = data

@st.cache_data
def load_auth_cached(): return load_json("auth", {})

def load_auth(): return load_auth_cached()

def save_auth(d): save_json("auth", d); load_auth_cached.clear()

def load_schools(): return load_json("schools", {})

def load_conf():
    default_conf = {"year": "6", "tournaments": DEFAULT_TOURNAMENTS, "limits": DEFAULT_LIMITS, "admin_password": "1234"}
    data = load_json("config", default_conf)
    if "limits" not in data: data["limits"] = DEFAULT_LIMITS
    if "tournaments" not in data: data["tournaments"] = DEFAULT_TOURNAMENTS
    return data

def save_conf(d): save_json("config", d)

def upload_file_to_gas(uploaded_file, school_name):
    if not GAS_WEBAPP_URL or GAS_WEBAPP_URL == "ここに貼り付け": return False, "GASのURLが設定されていません。"
    try:
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = os.path.splitext(uploaded_file.name)[1]; file_name = f"【{school_name}】_申込書_{timestamp}{ext}"
        base64_content = base64.b64encode(uploaded_file.getvalue()).decode('utf-8')
        payload = {"fileName": file_name, "mimeType": uploaded_file.type, "base64": base64_content}
        res_data = requests.post(GAS_WEBAPP_URL, json=payload).json()
        if res_data.get("status") == "success": return True, res_data.get("id")
        else: return False, res_data.get("message", "不明なエラー")
    except Exception as e: return False, str(e)

# ---------------------------------------------------------
# 4. ロジック 
# ---------------------------------------------------------
def create_backup():
    df = load_members_master(force_reload=False)
    ws_bk_mem = get_worksheet_safe(f"{V2_PREFIX}members_backup"); ws_bk_mem.clear()
    df_bk = df.fillna("")[MEMBERS_COLS]; ws_bk_mem.update([df_bk.columns.tolist()] + df_bk.astype(str).values.tolist())
    conf = load_conf(); ws_bk_conf = get_worksheet_safe(f"{V2_PREFIX}config_backup"); ws_bk_conf.update_acell('A1', json.dumps(conf, ensure_ascii=False))

def restore_from_backup():
    try:
        ws_bk_mem = get_worksheet_safe(f"{V2_PREFIX}members_backup"); recs = ws_bk_mem.get_all_records()
        df = pd.DataFrame(recs) if recs else pd.DataFrame(columns=MEMBERS_COLS)
        if not df.empty: df['grade'] = pd.to_numeric(df['grade'], errors='coerce').fillna(0).astype(int); save_members_master(df)
    except: return "名簿の復元に失敗しました"
    try:
        ws_bk_conf = get_worksheet_safe(f"{V2_PREFIX}config_backup"); val = ws_bk_conf.acell('A1').value
        if val: save_conf(json.loads(val))
    except: return "設定の復元に失敗しました"
    return "✅ バックアップから復元しました"

def perform_year_rollover():
    create_backup()
    if "v2_master_cache" in st.session_state: del st.session_state["v2_master_cache"]
    df = load_members_master(force_reload=True)
    if df.empty: return "データがありません"
    df['grade'] = df['grade'] + 1
    graduates = df[df['grade'] > 3].copy(); current = df[df['grade'] <= 3].copy()
    if not graduates.empty:
        auth = load_auth(); ws_grad = get_worksheet_safe(f"{V2_PREFIX}graduates")
        graduates["archived_school_name"] = graduates["school_id"].apply(lambda sid: auth.get(sid, {}).get("base_name", "不明"))
        graduates["archived_date"] = datetime.date.today().strftime("%Y-%m-%d")
        if not ws_grad.get_all_values(): ws_grad.append_row(graduates.columns.tolist())
        ws_grad.append_rows(graduates.astype(str).values.tolist())
    save_members_master(current)
    conf = load_conf()
    for tid in conf["tournaments"].keys(): save_entries(tid, {})
    try: conf["year"] = str(int(conf["year"]) + 1); save_conf(conf)
    except: pass
    return f"✅ 新年度更新完了。{len(graduates)}名の卒業生データをアーカイブしました。"

def get_merged_data(school_id, tournament_id):
    master = load_members_master(force_reload=False)
    if master.empty: return pd.DataFrame()
    my_members = master[master['school_id'] == school_id].copy()
    entries = load_entries(tournament_id, force_reload=False)
    cols_to_add = ["team_kata_chk", "team_kata_role", "team_kumi_chk", "team_kumi_role", "kata_chk", "kata_val", "kata_rank", "kumi_chk", "kumi_val", "kumi_rank"]
    for c in cols_to_add: my_members[f"last_{c}"] = my_members.apply(lambda r: entries.get(f"{r['school_id']}_{r['name']}", {}).get(c, None), axis=1)
    return my_members

def validate_counts(members_df, entries_data, limits, t_type, school_meta, school_id):
    errs = []
    for sex in ["男子", "女子"]:
        sex_df = members_df[members_df['sex'] == sex]
        cnt_tk = 0; cnt_tku = 0; cnt_ind_k_reg = 0; cnt_ind_k_sub = 0; cnt_ind_ku_reg = 0; cnt_ind_ku_sub = 0
        for _, r in sex_df.iterrows():
            uid = f"{school_id}_{r['name']}"; ent = entries_data.get(uid, {})
            if ent.get("team_kata_chk") and ent.get("team_kata_role") == "正": cnt_tk += 1
            if ent.get("team_kumi_chk") and ent.get("team_kumi_role") == "正": cnt_tku += 1
            if ent.get("kata_chk"):
                if ent.get("kata_val") == "補": cnt_ind_k_sub += 1
                elif ent.get("kata_val") == "正": cnt_ind_k_reg += 1 
            if ent.get("kumi_chk"):
                v = ent.get("kumi_val")
                if v == "補": cnt_ind_ku_sub += 1
                elif v == "正": cnt_ind_ku_reg += 1
                elif t_type != "standard" and v and v not in ["出場しない", "なし", "シード", "補"]: cnt_ind_ku_reg += 1
        if cnt_tk > 0:
            mn, mx = limits["team_kata"]["min"], limits["team_kata"]["max"]
            if not (mn <= cnt_tk <= mx): errs.append(f"❌ {sex}団体形: 正選手は {mn}～{mx}名で登録してください。(現在{cnt_tk}名)")
        if cnt_tku > 0:
            mode = school_meta.get("m_kumite_mode" if sex == "男子" else "w_kumite_mode", "none") if t_type == "shinjin" else "5"
            l_key = "team_kumite_3" if mode == "3" else "team_kumite_5"; mn, mx = limits[l_key]["min"], limits[l_key]["max"]
            if not (mn <= cnt_tku <= mx): errs.append(f"❌ {sex}団体組手({mode}人制): 正選手は {mn}～{mx}名で登録してください。(現在{cnt_tku}名)")
        if cnt_ind_k_reg > limits["ind_kata_reg"]["max"]: errs.append(f"❌ {sex}個人形(正): 上限 {limits['ind_kata_reg']['max']}名を超えています。(シード除く)")
        if cnt_ind_k_sub > limits["ind_kata_sub"]["max"]: errs.append(f"❌ {sex}個人形(補): 上限 {limits['ind_kata_sub']['max']}名を超えています。")
        if cnt_ind_ku_reg > limits["ind_kumi_reg"]["max"]: errs.append(f"❌ {sex}個人組手(正): 上限 {limits['ind_kumi_reg']['max']}名を超えています。(シード除く)")
        if cnt_ind_ku_sub > limits["ind_kumi_sub"]["max"]: errs.append(f"❌ {sex}個人組手(補): 上限 {limits['ind_kumi_sub']['max']}名を超えています。")
    return errs

# ---------------------------------------------------------
# 5. Excel生成
# ---------------------------------------------------------
def safe_write(ws, target, value, align_center=False):
    if value is None: value = ""
    cell = ws[target] if isinstance(target, str) else ws.cell(row=target[0], column=target[1])
    if isinstance(cell, MergedCell):
        for r in ws.merged_cells.ranges:
            if cell.coordinate in r: cell = ws.cell(row=r.min_row, column=r.min_col); break
    val_str = str(value)
    if val_str.endswith("年") and val_str[:-1].isdigit(): val_str = val_str.replace("年", "")
    cell.value = val_str
    if align_center: cell.alignment = Alignment(horizontal='center', vertical='center')

def generate_excel(school_id, school_data, members_df, t_id, t_conf):
    coords = COORD_DEF; template_file = t_conf.get("template", "template.xlsx")
    try: wb = openpyxl.load_workbook(template_file); ws = wb.active
    except: return None, f"{template_file} が見つかりません。"
    conf = load_conf(); safe_write(ws, coords["year"], conf.get("year", "")); safe_write(ws, coords["tournament_name"], t_conf.get("name", ""))
    safe_write(ws, coords["date"], f"令和{datetime.date.today().year-2018}年{datetime.date.today().month}月{datetime.date.today().day}日")
    bn = school_data.get("base_name", ""); safe_write(ws, coords["school_name"], bn); safe_write(ws, coords["principal"], school_data.get("principal", ""))
    advs = school_data.get("advisors", []); safe_write(ws, coords["head_advisor"], advs[0]["name"] if advs else "")
    for i, a in enumerate(advs[:4]):
        c = coords["advisors"][i]; safe_write(ws, c["name"], a["name"]); safe_write(ws, c["d1"], "○" if a.get("d1") else "×", True); safe_write(ws, c["d2"], "○" if a.get("d2") else "×", True)
    cols = coords["cols"]; members_df['sex_rank'] = members_df['sex'].map({'男子': 0, '女子': 1}); members_df['grade_rank'] = members_df['grade'].map({3: 0, 2: 1, 1: 2})
    def get_sort_key(row):
        try: return float(row['display_order']) if pd.notna(row['display_order']) and str(row['display_order']).strip() else 999999.0
        except: return 999999.0
    members_df['custom_order'] = members_df.apply(get_sort_key, axis=1)
    target_grades = [int(g) for g in t_conf.get('grades', [1, 2, 3])]
    entries = members_df[members_df['grade'].isin(target_grades)].sort_values(by=['custom_order', 'sex_rank', 'grade_rank', 'name'])
    for i, (_, row) in enumerate(entries.iterrows()):
        r = coords["start_row"] + (i // coords["cap"] * coords["offset"]) + (i % coords["cap"])
        safe_write(ws, (r, cols["name"]), row["name"]); safe_write(ws, (r, cols["grade"]), row["grade"]); safe_write(ws, (r, cols["dob"]), row["dob"]); safe_write(ws, (r, cols["jkf_no"]), row["jkf_no"])
        sex = row["sex"]; tk_c = cols["m_team_kata"] if sex=="男子" else cols["w_team_kata"]; tku_c = cols["m_team_kumite"] if sex=="男子" else cols["w_team_kumite"]
        if row.get("last_team_kata_chk"): safe_write(ws, (r, tk_c), "補" if row.get("last_team_kata_role")=="補" else "○", True)
        if row.get("last_team_kumi_chk"): safe_write(ws, (r, tku_c), "補" if row.get("last_team_kumi_role")=="補" else "○", True)
        k_c = cols["m_kata"] if sex=="男子" else cols["w_kata"]; ku_c = cols["m_kumite"] if sex=="男子" else cols["w_kumite"]
        if row.get("last_kata_chk"):
            v, rk = row.get("last_kata_val"), row.get("last_kata_rank", "")
            txt = "補" if v=="補" else (f"シ{rk}" if v=="シード" else f"○{rk}")
            safe_write(ws, (r, k_c), txt, True)
        if row.get("last_kumi_chk"):
            v, rk = row.get("last_kumi_val"), row.get("last_kumi_rank", "")
            if v=="補": txt = "補"
            elif t_conf["type"]=="standard": txt = f"シ{rk}" if v=="シード" else f"○{rk}"
            else: txt = str(v)
            safe_write(ws, (r, ku_c), txt, True)
    fname = f"申込書_{bn}.xlsx"; wb.save(fname); return fname, "成功"

# ★大幅改修: 個人戦・団体戦の統合出力対応
def generate_tournament_excel(all_data, t_type, auth_data, target_sex, target_category, template_file):
    sheets_data = {}
    processed_schools = set() # 団体戦の重複登録防止用
    
    for row in all_data:
        name, sid, sex = row['name'], row['school_id'], row['sex']
        
        if sex != target_sex:
            continue

        s_data = auth_data.get(sid, {})
        school_short = s_data.get("short_name", s_data.get("base_name", ""))
        
        # --- 個人形 ---
        if target_category == "形" and row.get('kata_chk'):
            k_val = row.get('kata_val')
            k_rank = row.get('kata_rank', '')
            if k_val and k_val not in ['補', 'なし', '出場しない']:
                sn = f"{sex}個人形"
                rank_cell = k_rank if k_val == '正' else ''
                seed_cell = k_rank if k_val == 'シード' else ''
                if sn not in sheets_data: sheets_data[sn] = []
                sheets_data[sn].append([rank_cell, name, school_short, seed_cell])
                
        # --- 個人組手 ---
        elif target_category == "組手" and row.get('kumi_chk'):
            ku_val = row.get('kumi_val')
            ku_rank = row.get('kumi_rank', '')
            if ku_val and ku_val not in ['補', 'なし', '出場しない']:
                if t_type == 'standard':
                    sn = f"{sex}個人組手"
                    is_seed = (ku_val == 'シード')
                    is_reg = (ku_val == '正')
                else:
                    sn = f"{sex}個人組手_{ku_val}"
                    is_seed = False
                    is_reg = True 
                rank_cell = ku_rank if is_reg else ''
                seed_cell = ku_rank if is_seed else ''
                if sn not in sheets_data: sheets_data[sn] = []
                sheets_data[sn].append([rank_cell, name, school_short, seed_cell])

        # --- 団体形 ---
        elif target_category == "団体形" and row.get('team_kata_chk'):
            if sid not in processed_schools:
                sn = f"{sex}団体形"
                if sn not in sheets_data: sheets_data[sn] = []
                # 名前欄に学校名を入れることで、VBAのエラーを回避
                sheets_data[sn].append(["", school_short, school_short, ""]) 
                processed_schools.add(sid)

        # --- 団体組手 ---
        elif target_category == "団体組手" and row.get('team_kumi_chk'):
            if sid not in processed_schools:
                sn = f"{sex}団体組手"
                if sn not in sheets_data: sheets_data[sn] = []
                # 名前欄に学校名を入れることで、VBAのエラーを回避
                sheets_data[sn].append(["", school_short, school_short, ""]) 
                processed_schools.add(sid)

    output = io.BytesIO()
    has_template = os.path.exists(template_file)
    
    if has_template:
        wb = openpyxl.load_workbook(template_file, keep_vba=True)
        mime_type = "application/vnd.ms-excel.sheet.macroEnabled.12"
        ext = "xlsm"
    else:
        wb = openpyxl.Workbook()
        if wb.sheetnames: wb.remove(wb.active)
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    # データを流し込む
    for s_name in sorted(sheets_data.keys()):
        recs = sheets_data[s_name]
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(s_name)
        
        ws.append(["順位", "名前", "学校名", "シード枠"])
        for rec in recs:
            ws.append(rec)
            
    if not wb.sheetnames: wb.create_sheet("データなし")
        
    wb.save(output)
    return output.getvalue(), mime_type, ext


def generate_summary_excel(master_df, entries, auth_data, t_type):
    rows = []
    for s_id, s_data in sorted(auth_data.items(), key=lambda x: to_safe_int(x[1].get('school_no'))):
        s_name = s_data.get("short_name", s_data.get("base_name", "")); s_members = master_df[master_df['school_id'] == s_id]
        m_tk, m_tku, w_tk, w_tku, m_k, m_ku, w_k, w_ku, regs = "", "", "", "", 0, 0, 0, 0, set()
        for _, r in s_members.iterrows():
            ent = entries.get(f"{s_id}_{r['name']}", {}); sex = r['sex']
            if sex == "男子":
                if ent.get("team_kata_chk"): m_tk = "○"
                if ent.get("team_kumi_chk"): m_tku = "○"
                if ent.get("kata_chk") and ent.get("kata_val") not in ["補","なし","出場しない"]: m_k += 1
                if ent.get("kumi_chk") and ent.get("kumi_val") not in ["補","なし","出場しない"]: m_ku += 1
            else:
                if ent.get("team_kata_chk"): w_tk = "○"
                if ent.get("team_kumi_chk"): w_tku = "○"
                if ent.get("kata_chk") and ent.get("kata_val") not in ["補","なし","出場しない"]: w_k += 1
                if ent.get("kumi_chk") and ent.get("kumi_val") not in ["補","なし","出場しない"]: w_ku += 1
            if (ent.get("team_kata_chk") and ent.get("team_kata_role")=="正") or (ent.get("team_kumi_chk") and ent.get("team_kumi_role")=="正") or (ent.get("kata_chk") and ent.get("kata_val") not in ["補","なし","出場しない"]) or (ent.get("kumi_chk") and ent.get("kumi_val") not in ["補","なし","出場しない"]): regs.add(r['name'])
        rows.append({"学校No": s_data.get('school_no',''), "学校名": s_name, "男団体形": m_tk, "男団体組手": m_tku, "男個人形": m_k if m_k>0 else "", "男個人組手": m_ku if m_ku>0 else "", "女団体形": w_tk, "女団体組手": w_tku, "女個人形": w_k if w_k>0 else "", "女個人組手": w_ku if w_ku>0 else "", "正選手合計": len(regs)})
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer: pd.DataFrame(rows).to_excel(writer, sheet_name="参加校一覧", index=False)
    return output.getvalue()

def generate_advisor_excel(schools_data, auth_data):
    rows = []
    for s_id, s_auth in sorted(auth_data.items(), key=lambda x: to_safe_int(x[1].get('school_no'))):
        s_name = s_auth.get("short_name", s_auth.get("base_name", ""))
        for a in s_auth.get("advisors", []):
            if a.get("name"): rows.append({"No": s_auth.get('school_no',''), "学校名": s_name, "顧問氏名": a["name"], "役割": a.get("role","審判"), "1日目": "○" if a.get("d1") else "×", "2日目": "○" if a.get("d2") else "×"})
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer: pd.DataFrame(rows).to_excel(writer, sheet_name="顧問一覧", index=False)
    return output.getvalue()

# ---------------------------------------------------------
# 7. UI 
# ---------------------------------------------------------
def school_page(s_id):
    st.markdown("""<style>div[data-testid="stRadio"] > div { flex-direction: row; }</style>""", unsafe_allow_html=True)
    auth = load_auth(); s_data = auth.get(s_id, {}); base_name = s_data.get("base_name", "")
    col_h1, col_h2 = st.columns([3, 1])
    with col_h1: st.markdown(f"### {base_name}高等学校")
    with col_h2:
        if st.button("🚪 ログアウト", type="secondary", use_container_width=True): st.query_params.clear(); st.session_state.clear(); st.rerun()
    st.divider()
    conf = load_conf(); active_tid = next((k for k, v in conf["tournaments"].items() if v["active"]), None)
    if not active_tid: st.error("現在受付中の大会はありません。"); return
    t_conf = conf["tournaments"][active_tid]; st.markdown(f"## 🥋 **令和{conf.get('year','〇')}年度 {t_conf['name']}** <small>エントリー画面</small>", unsafe_allow_html=True)
    
    if st.button("🔄 データを最新にする"):
        if "v2_master_cache" in st.session_state: del st.session_state["v2_master_cache"]
        if f"v2_entry_cache_{active_tid}" in st.session_state: del st.session_state[f"v2_entry_cache_{active_tid}"]
        st.success("最新データを読み込みました"); time.sleep(0.5); st.rerun()

    menu = ["① 顧問登録", "② 部員名簿登録", "③ 大会エントリー"]
    if "school_menu_idx" not in st.session_state: st.session_state["school_menu_idx"] = 0
    selected_view = st.radio("メニュー選択", menu, index=st.session_state["school_menu_idx"], key="school_menu_radio", horizontal=True, label_visibility="collapsed")
    st.session_state["school_menu_idx"] = menu.index(selected_view)
    st.markdown("---")

    if selected_view == "① 顧問登録":
        st.warning("⚠️ **重要:** 編集内容は自動保存されません。変更後は必ず下の **『💾 顧問情報を保存』** ボタンを押してください。")
        np = st.text_input("校長名", s_data.get("principal", ""))
        adv_df = pd.DataFrame(s_data.get("advisors", []))
        for c in ["name", "role", "d1", "d2"]:
            if c not in adv_df.columns: adv_df[c] = ""
        edited_adv_df = st.data_editor(adv_df[["name", "role", "d1", "d2"]], column_config={"name": "氏名", "role": st.column_config.SelectboxColumn("役割", options=["審判", "競技記録", "係員"], required=True), "d1": "1日目", "d2": "2日目"}, num_rows="dynamic", use_container_width=True, hide_index=True)
        if st.button("💾 顧問情報を保存", type="primary"):
            if edited_adv_df["name"].isnull().any() or (edited_adv_df["name"] == "").any(): st.error("❌ 氏名が未入力です"); return
            with st.spinner("保存中..."):
                load_auth_cached.clear(); latest_auth = load_auth(); cur_s = latest_auth.get(s_id, s_data)
                cur_s["principal"], cur_s["advisors"] = np, edited_adv_df.to_dict(orient="records")
                latest_auth[s_id] = cur_s; save_auth(latest_auth); st.success("✅ 保存完了"); time.sleep(1); st.rerun()

    elif selected_view == "② 部員名簿登録":
        st.warning("⚠️ **重要:** 編集内容は自動保存されません。変更後は必ず下の **『💾 名簿を保存して更新』** ボタンを押してください。")
        st.caption("💡 **表示順について:** 「No.」列に数字を入力すると、その順番に表示されます。指定がない場合は学年・名前順になります。")
        master = load_members_master(force_reload=False); my_m = master[master['school_id']==s_id].copy()
        disp_df = my_m[["display_order", "name", "sex", "grade", "dob", "jkf_no"]].copy()
        edited_mem_df = st.data_editor(disp_df, column_config={"display_order": st.column_config.NumberColumn("No.", step=1), "name": "氏名", "sex": st.column_config.SelectboxColumn("性別", options=["男子", "女子"], required=True), "grade": st.column_config.SelectboxColumn("学年", options=[1, 2, 3], required=True), "dob": "生年月日", "jkf_no": "JKF番号"}, num_rows="dynamic", use_container_width=True, hide_index=True)
        if st.button("💾 名簿を保存して更新", type="primary"):
            with st.spinner("保存中..."):
                create_backup(); edited_mem_df["school_id"], edited_mem_df["active"] = s_id, True
                edited_mem_df['display_order'] = edited_mem_df['display_order'].apply(lambda x: str(int(x)) if pd.notnull(x) and str(x).strip() != "" else "")
                for c in MEMBERS_COLS:
                    if c not in edited_mem_df.columns: edited_mem_df[c] = ""
                latest_master = load_members_master(force_reload=True); new_master = pd.concat([latest_master[latest_master['school_id'] != s_id], edited_mem_df[MEMBERS_COLS]], ignore_index=True)
                save_members_master(new_master); st.success("✅ 更新完了"); time.sleep(1); st.rerun()
        st.divider()
        st.markdown("##### 📋 登録済み部員リスト（確認用）")
        master_check = load_members_master(force_reload=False)
        my_check = master_check[master_check['school_id']==s_id]
        def sort_key(row):
            try: return float(row['display_order']) if pd.notna(row['display_order']) and str(row['display_order']).strip() else 999999.0
            except: return 999999.0
        my_check['sort_k'] = my_check.apply(sort_key, axis=1)
        c_male, c_female = st.columns(2)
        with c_male:
            st.markdown("###### 🚹 男子部員")
            m_df = my_check[my_check['sex'] == '男子'].sort_values(by=['sort_k', 'grade', 'name'], ascending=[True, False, True])
            if not m_df.empty: st.dataframe(m_df[['display_order', 'grade','name','jkf_no']].rename(columns={'display_order':'No.','grade':'学年','name':'氏名','jkf_no':'JKF番号'}), hide_index=True, use_container_width=True)
            else: st.caption("登録なし")
        with c_female:
            st.markdown("###### 🚺 女子部員")
            w_df = my_check[my_check['sex'] == '女子'].sort_values(by=['sort_k', 'grade', 'name'], ascending=[True, False, True])
            if not w_df.empty: st.dataframe(w_df[['display_order', 'grade','name','jkf_no']].rename(columns={'display_order':'No.','grade':'学年','name':'氏名','jkf_no':'JKF番号'}), hide_index=True, use_container_width=True)
            else: st.caption("登録なし")

    elif selected_view == "③ 大会エントリー":
        target_grades = [int(g) for g in t_conf['grades']]
        st.markdown(f"**出場対象学年:** {target_grades} 年生")
        st.markdown("""<div style="background-color:#ffebee; border:1px solid #ef9a9a; padding:10px; border-radius:5px; color:#c62828;"><h4 style="margin:0;">⚠️ 順位入力について</h4><p style="font-weight:bold; margin:5px 0;">シード権やトーナメント配置の優先順位に使用します。補欠の場合は入力不要です。（例：1, 2, 3...）</p></div><br>""", unsafe_allow_html=True)
        merged = get_merged_data(s_id, active_tid)
        if merged.empty: st.warning("名簿を登録してください。"); return
        merged['sex_rank'] = merged['sex'].map({'男子': 0, '女子': 1}); merged['grade_rank'] = merged['grade'].map({3: 0, 2: 1, 1: 2})
        valid_members = merged[merged['grade'].isin(target_grades)].copy()
        def get_sort_key_ent(row):
            try: return float(row['display_order']) if pd.notna(row['display_order']) and str(row['display_order']).strip() else 999999.0
            except: return 999999.0
        valid_members['custom_order'] = valid_members.apply(get_sort_key_ent, axis=1)
        valid_members = valid_members.sort_values(by=['custom_order', 'sex_rank', 'grade_rank', 'name'])
        
        entries_update = load_entries(active_tid, force_reload=False); school_meta = entries_update.get(f"_meta_{s_id}", {"m_kumite_mode": "none", "w_kumite_mode": "none"})
        m_mode, w_mode = "5", "5"
        if t_conf["type"] == "shinjin":
            with st.expander("団体組手の設定 (新人戦)", expanded=True):
                c_m, c_w = st.columns(2)
                cur_m = school_meta.get("m_kumite_mode", "none"); idx_m = ["none", "5", "3"].index(cur_m) if cur_m in ["none", "5", "3"] else 0
                new_m = c_m.radio("男子 団体組手", ["出場しない", "5人制", "3人制"], index=idx_m, horizontal=True)
                m_mode = "none" if new_m == "出場しない" else ("5" if new_m == "5人制" else "3")
                cur_w = school_meta.get("w_kumite_mode", "none"); idx_w = ["none", "5", "3"].index(cur_w) if cur_w in ["none", "5", "3"] else 0
                new_w = c_w.radio("女子 団体組手", ["出場しない", "5人制", "3人制"], index=idx_w, horizontal=True)
                w_mode = "none" if new_w == "出場しない" else ("5" if new_w == "5人制" else "3")
                if new_m != cur_m or new_w != cur_w:
                    school_meta["m_kumite_mode"] = m_mode; school_meta["w_kumite_mode"] = w_mode; entries_update[f"_meta_{s_id}"] = school_meta; save_entries(active_tid, entries_update)

        with st.form("entry_form_unified"):
            cols = st.columns([1.7, 1.4, 1.4, 0.1, 3.1, 3.1]); cols[0].markdown("**氏名**"); cols[1].markdown("**団体形**"); cols[2].markdown("**団体組手**"); cols[4].markdown("**個人形**"); cols[5].markdown("**個人組手**")
            form_buffer = {}
            for i, r in valid_members.iterrows():
                uid = f"{s_id}_{r['name']}"; ns = 'background-color:#e8f5e9; padding:2px; font-weight:bold;' if r['sex']=="男子" else 'background-color:#ffebee; padding:2px; font-weight:bold;'
                c = st.columns([1.7, 1.4, 1.4, 0.1, 3.1, 3.1])
                c[0].markdown(f'<span style="{ns}">{r["grade"]}年 {r["name"]}</span>', unsafe_allow_html=True)
                val_tk = c[1].radio(f"tk_{uid}", ["なし", "正", "補"], index=["なし", "正", "補"].index(r.get("last_team_kata_role", "なし") if r.get("last_team_kata_role") in ["なし", "正", "補"] else "なし"), horizontal=True, label_visibility="collapsed")
                val_tku = c[2].radio(f"tku_{uid}", ["なし", "正", "補"], index=["なし", "正", "補"].index(r.get("last_team_kumi_role", "なし") if r.get("last_team_kumi_role") in ["なし", "正", "補"] else "なし"), horizontal=True, label_visibility="collapsed") if (m_mode if r['sex']=="男子" else w_mode) != "none" else "なし"
                if (m_mode if r['sex']=="男子" else w_mode) == "none": c[2].caption("-")
                opts_k = ["なし", "正", "補", "シード"] if t_conf["type"]=="standard" else ["なし", "正", "補"]
                ck1, ck2 = c[4].columns([1.5, 1])
                val_k = ck1.radio(f"k_{uid}", opts_k, index=opts_k.index(r.get("last_kata_val","なし") if r.get("last_kata_val") in opts_k else "なし"), horizontal=True, label_visibility="collapsed")
                rk_k = ck2.text_input("順位", r.get("last_kata_rank",""), key=f"rk_k_{uid}", label_visibility="collapsed", placeholder="順位")
                c5a, c5b = c[5].columns([1.8, 1])
                w_list = ["出場しない"] + [f"{w.strip()}kg級" for w in t_conf.get("weights_m" if r['sex']=="男子" else "weights_w", "").split(",")] + ["補欠"]
                if t_conf["type"]=="standard":
                    ku_v = c5a.radio(f"ku_{uid}", ["なし", "正", "補", "シード"], index=["なし", "正", "補", "シード"].index(r.get("last_kumi_val","なし") if r.get("last_kumi_val") in ["なし", "正", "補", "シード"] else "なし"), horizontal=True, label_visibility="collapsed")
                else:
                    cur_ku = r.get("last_kumi_val", "出場しない"); idx_ku = w_list.index(cur_ku) if cur_ku in w_list else 0
                    ku_v = c5a.selectbox("階級", w_list, index=idx_ku, key=f"sel_ku_{uid}", label_visibility="collapsed")
                rk_ku = c5b.text_input("順位", r.get("last_kumi_rank",""), key=f"rk_ku_{uid}", label_visibility="collapsed", placeholder="順位")
                form_buffer[uid] = {"val_tk": val_tk, "val_tku": val_tku, "val_k": val_k, "rank_k": rk_k, "ku_val": ku_v, "rank_ku": rk_ku, "name": r["name"], "sex": r["sex"]}
            
            if st.form_submit_button("✅ エントリーを保存 (全員分)"):
                has_error = False; temp_processed = {}; duplicate_checker = {} 
                for uid, raw in form_buffer.items():
                    k_chk = (raw["val_k"] != "なし"); k_val = raw["val_k"] if k_chk else ""
                    if (k_val == "補" or k_val == "なし") and raw["rank_k"]:
                        st.error(f"❌ {raw['name']} 個人形: 「{k_val}」ですが順位が入力されています。順位を削除してください。"); has_error = True
                    if k_chk:
                        if (k_val == "正" or k_val == "シード") and not raw["rank_k"]:
                            st.error(f"❌ {raw['name']} 個人形: {k_val}選手の実績順位が入力されていません。"); has_error = True
                        if (k_val == "正" or k_val == "シード") and raw["rank_k"]:
                            check_key = f"{raw['sex']}_kata_{k_val}"; clean_rank = to_half_width(raw["rank_k"])
                            if check_key not in duplicate_checker: duplicate_checker[check_key] = {}
                            if clean_rank not in duplicate_checker[check_key]: duplicate_checker[check_key][clean_rank] = []
                            duplicate_checker[check_key][clean_rank].append(raw["name"])

                    ku_chk = (raw["ku_val"] not in ["なし", "出場しない"]); ku_val = raw["ku_val"] if ku_chk else ""
                    if (ku_val in ["補", "なし", "出場しない", "補欠"]) and raw["rank_ku"]:
                        st.error(f"❌ {raw['name']} 個人組手: 「{ku_val}」ですが順位が入力されています。順位を削除してください。"); has_error = True
                    if ku_chk:
                        is_reg = (t_conf["type"] == "weight" and ku_val != "補欠") or (t_conf["type"] == "standard" and ku_val == "正")
                        is_seed = (t_conf["type"] == "standard" and ku_val == "シード")
                        if (is_reg or is_seed) and not raw["rank_ku"]:
                            st.error(f"❌ {raw['name']} 個人組手: 実績順位が入力されていません。"); has_error = True
                        if t_conf["type"] == "standard" and (is_reg or is_seed) and raw["rank_ku"]:
                            check_key = f"{raw['sex']}_kumite_{'シード' if is_seed else '正'}"; clean_rank = to_half_width(raw["rank_ku"])
                            if check_key not in duplicate_checker: duplicate_checker[check_key] = {}
                            if clean_rank not in duplicate_checker[check_key]: duplicate_checker[check_key][clean_rank] = []
                            duplicate_checker[check_key][clean_rank].append(raw["name"])

                    temp_processed[uid] = {"team_kata_chk": raw["val_tk"]!="なし", "team_kata_role": raw["val_tk"] if raw["val_tk"]!="なし" else "", "team_kumi_chk": raw["val_tku"]!="なし", "team_kumi_role": raw["val_tku"] if raw["val_tku"]!="なし" else "", "kata_chk": k_chk, "kata_val": k_val, "kata_rank": to_half_width(raw["rank_k"]), "kumi_chk": ku_chk, "kumi_val": ku_val, "kumi_rank": to_half_width(raw["rank_ku"])}
                
                for key, ranks in duplicate_checker.items():
                    for rank_val, names in ranks.items():
                        if len(names) > 1:
                            parts = key.split("_"); st.error(f"❌ {parts[0]} 個人{'形' if parts[1]=='kata' else '組手'} ({parts[2]}選手) で順位『{rank_val}』が重複しています: {', '.join(names)}"); has_error = True

                if not has_error:
                    with st.spinner("💾 エントリーを保存しています..."):
                        cur_entries = load_entries(active_tid, force_reload=True); cur_entries.update(temp_processed)
                        errs = validate_counts(valid_members, cur_entries, conf["limits"], t_conf["type"], {"m_kumite_mode":m_mode, "w_kumite_mode":w_mode}, s_id)
                        if errs:
                            for e in errs: st.error(e)
                        else: save_entries(active_tid, cur_entries); st.success("✅ 保存しました！"); time.sleep(2); st.rerun()

        st.markdown("---")
        st.markdown("#### 📥 申込書の出力と提出")
        st.info("出力したExcelファイルに公印を押し、PDFや画像にしてから右側の枠へ提出してください。")
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("##### 1. 申込書の作成")
            if st.button("📄 Excel申込書を作成する", type="secondary", use_container_width=True):
                 final_m = get_merged_data(s_id, active_tid); fp, msg = generate_excel(s_id, s_data, final_m, active_tid, t_conf)
                 if fp:
                     with open(fp, "rb") as f: st.download_button("📥 ダウンロード", f, fp, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with c2:
            st.markdown("##### 2. 申込書のアップロード")
            u_file = st.file_uploader("ファイルを選択 (PDF, JPG, PNG 等)", type=['pdf', 'jpg', 'jpeg', 'png'], label_visibility="collapsed")
            if u_file:
                if st.button("✅ 申込書を提出する", type="primary", use_container_width=True):
                    with st.spinner("安全に送信中..."):
                        ok, res = upload_file_to_gas(u_file, base_name)
                        if ok: st.success("🎉 提出完了しました！管理者が確認いたします。")
                        else: st.error(f"❌ 失敗: {res}")

def admin_page():
    st.title("🔧 管理者画面")
    conf = load_conf()
    if not st.session_state.get("admin_ok", False):
        pw = st.text_input("Admin Password", type="password")
        if st.button("ログイン"):
            if pw == conf.get("admin_password", "1234"): st.session_state["admin_ok"] = True; st.rerun()
            else: st.error("パスワードが違います")
        return
    auth = load_auth(); admin_menu = ["🏆 大会設定", "📥 データ出力", "🏫 アカウント", "📅 年次処理"]
    admin_tab = st.radio("メニュー", admin_menu, index=st.session_state.get("admin_menu_idx", 0), horizontal=True)
    st.session_state["admin_menu_idx"] = admin_menu.index(admin_tab)
    st.divider()

    if admin_tab == "📥 データ出力":
        st.subheader("大会データの出力")
        tid = next((k for k, v in conf["tournaments"].items() if v["active"]), "kantou")
        if st.button("🔄 最新データで集計を開始"):
            with st.spinner("集計中..."):
                master = load_members_master(force_reload=True); entries = load_entries(tid, force_reload=True); full_data = []
                for _, m in master.iterrows():
                    ent = entries.get(f"{m['school_id']}_{m['name']}", {})
                    if ent.get("kata_chk") or ent.get("kumi_chk") or ent.get("team_kata_chk") or ent.get("team_kumi_chk"):
                        row = m.to_dict(); row.update(ent); full_data.append(row)
                
                # --- 個人戦 4種目 ---
                tm_m, mime_m, ext_m = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "男子", "組手", "template_kumite_m.xlsm")
                st.session_state["xlsx_kumite_m"] = tm_m; st.session_state["mime_kumite_m"] = mime_m; st.session_state["ext_kumite_m"] = ext_m
                
                tm_w, mime_w, ext_w = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "女子", "組手", "template_kumite_w.xlsm")
                st.session_state["xlsx_kumite_w"] = tm_w; st.session_state["mime_kumite_w"] = mime_w; st.session_state["ext_kumite_w"] = ext_w
                
                tk_m, mime_km, ext_km = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "男子", "形", "template_kata_m.xlsm")
                st.session_state["xlsx_kata_m"] = tk_m; st.session_state["mime_kata_m"] = mime_km; st.session_state["ext_kata_m"] = ext_km
                
                tk_w, mime_kw, ext_kw = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "女子", "形", "template_kata_w.xlsm")
                st.session_state["xlsx_kata_w"] = tk_w; st.session_state["mime_kata_w"] = mime_kw; st.session_state["ext_kata_w"] = ext_kw
                
                # --- 団体戦 4種目 ---
                tt_km_m, mime_tkm_m, ext_tkm_m = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "男子", "団体組手", "template_team_kumite_m.xlsm")
                st.session_state["xlsx_team_kumite_m"] = tt_km_m; st.session_state["mime_team_kumite_m"] = mime_tkm_m; st.session_state["ext_team_kumite_m"] = ext_tkm_m
                
                tt_km_w, mime_tkm_w, ext_tkm_w = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "女子", "団体組手", "template_team_kumite_w.xlsm")
                st.session_state["xlsx_team_kumite_w"] = tt_km_w; st.session_state["mime_team_kumite_w"] = mime_tkm_w; st.session_state["ext_team_kumite_w"] = ext_tkm_w
                
                tt_kt_m, mime_tkt_m, ext_tkt_m = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "男子", "団体形", "template_team_kata_m.xlsm")
                st.session_state["xlsx_team_kata_m"] = tt_kt_m; st.session_state["mime_team_kata_m"] = mime_tkt_m; st.session_state["ext_team_kata_m"] = ext_tkt_m
                
                tt_kt_w, mime_tkt_w, ext_tkt_w = generate_tournament_excel(full_data, conf["tournaments"][tid]["type"], auth, "女子", "団体形", "template_team_kata_w.xlsm")
                st.session_state["xlsx_team_kata_w"] = tt_kt_w; st.session_state["mime_team_kata_w"] = mime_tkt_w; st.session_state["ext_team_kata_w"] = ext_tkt_w
                
                st.session_state["xlsx_summ"] = generate_summary_excel(master, entries, auth, conf["tournaments"][tid]["type"])
                st.session_state["xlsx_adv"] = generate_advisor_excel(load_schools(), auth)
                st.session_state["xlsx_ts"] = datetime.datetime.now().strftime("%H:%M:%S")
                
        if "xlsx_ts" in st.session_state:
            st.success(f"✅ 集計完了 ({st.session_state['xlsx_ts']})")
            
            st.markdown("##### 📥 個人戦 トーナメント作成データ")
            col1, col2, col3, col4 = st.columns(4)
            col1.download_button("🥋 男子個人組手", st.session_state["xlsx_kumite_m"], f"男子個人組手.{st.session_state['ext_kumite_m']}", mime=st.session_state["mime_kumite_m"])
            col2.download_button("🥋 女子個人組手", st.session_state["xlsx_kumite_w"], f"女子個人組手.{st.session_state['ext_kumite_w']}", mime=st.session_state["mime_kumite_w"])
            col3.download_button("🥋 男子個人形", st.session_state["xlsx_kata_m"], f"男子個人形.{st.session_state['ext_kata_m']}", mime=st.session_state["mime_kata_m"])
            col4.download_button("🥋 女子個人形", st.session_state["xlsx_kata_w"], f"女子個人形.{st.session_state['ext_kata_w']}", mime=st.session_state["mime_kata_w"])
            
            st.markdown("##### 📥 団体戦 トーナメント作成データ")
            col5, col6, col7, col8 = st.columns(4)
            col5.download_button("👥 男子団体組手", st.session_state["xlsx_team_kumite_m"], f"男子団体組手.{st.session_state['ext_team_kumite_m']}", mime=st.session_state["mime_team_kumite_m"])
            col6.download_button("👥 女子団体組手", st.session_state["xlsx_team_kumite_w"], f"女子団体組手.{st.session_state['ext_team_kumite_w']}", mime=st.session_state["mime_team_kumite_w"])
            col7.download_button("👥 男子団体形", st.session_state["xlsx_team_kata_m"], f"男子団体形.{st.session_state['ext_team_kata_m']}", mime=st.session_state["mime_team_kata_m"])
            col8.download_button("👥 女子団体形", st.session_state["xlsx_team_kata_w"], f"女子団体形.{st.session_state['ext_team_kata_w']}", mime=st.session_state["mime_team_kata_w"])
            
            st.markdown("##### 📊 その他データ")
            c1, c2 = st.columns(2)
            c1.download_button("📊 参加校一覧集計", st.session_state["xlsx_summ"], "summary.xlsx")
            c2.download_button("👔 顧問リスト", st.session_state["xlsx_adv"], "advisors.xlsx")
            
    elif admin_tab == "🏆 大会設定":
        st.subheader("基本設定")
        with st.form("conf_basic"):
            new_year = st.text_input("現在の年度", conf.get("year", "6"))
            t_opts = list(conf["tournaments"].keys())
            active_now = next((k for k, v in conf["tournaments"].items() if v["active"]), None)
            new_active = st.radio("受付中の大会", t_opts, index=t_opts.index(active_now) if active_now else 0, format_func=lambda x: conf["tournaments"][x]["name"])
            if st.form_submit_button("設定を保存 & 大会切替"):
                conf["year"] = new_year
                if new_active != active_now:
                    for k in conf["tournaments"]: conf["tournaments"][k]["active"] = (k == new_active)
                save_conf(conf); st.success("保存しました"); time.sleep(0.5); st.rerun()
        st.divider()
        with st.expander("参加人数制限の設定", expanded=True):
            with st.form("conf_limits"):
                lm = conf["limits"]; c1, c2 = st.columns(2)
                lm["team_kata"]["min"] = c1.number_input("団体形 下限", 0, 10, lm["team_kata"]["min"]); lm["team_kata"]["max"] = c2.number_input("団体形 上限", 0, 10, lm["team_kata"]["max"])
                c1, c2 = st.columns(2)
                lm["team_kumite_5"]["min"] = c1.number_input("団体組手(5人) 下限", 0, 10, lm["team_kumite_5"]["min"]); lm["team_kumite_5"]["max"] = c2.number_input("団体組手(5人) 上限", 0, 10, lm["team_kumite_5"]["max"])
                st.caption("個人戦 (上限のみ)"); c1, c2 = st.columns(2)
                lm["ind_kata_reg"]["max"] = c1.number_input("個人形(正) 上限", 0, 10, lm["ind_kata_reg"]["max"]); lm["ind_kata_sub"]["max"] = c2.number_input("個人形(補) 上限", 0, 10, lm["ind_kata_sub"]["max"])
                if st.form_submit_button("人数制限を保存"): conf["limits"] = lm; save_conf(conf); st.success("保存しました")
        with st.expander("🔐 管理者パスワード変更"):
            with st.form("admin_pw_change"):
                new_pw = st.text_input("新しい管理者パスワード", type="password")
                if st.form_submit_button("パスワードを変更して保存"):
                    if len(new_pw) >= 4: conf["admin_password"] = new_pw; save_conf(conf); st.success("変更しました"); time.sleep(1); st.session_state["admin_ok"] = False; st.rerun()
                    else: st.error("4文字以上にしてください")

    elif admin_tab == "🏫 アカウント":
        st.subheader("アカウント管理")
        recs = []
        for sid, d in auth.items(): recs.append({"ID": sid, "基本名": d.get("base_name",""), "略称": d.get("short_name", d.get("base_name","")), "No": d.get("school_no", 999), "Password": d.get("password",""), "校長名": d.get("principal","")})
        edited = st.data_editor(pd.DataFrame(recs), disabled=["ID"])
        if st.button("変更を保存"):
            for _, row in edited.iterrows():
                sid = row["ID"]
                if sid in auth: auth[sid].update({"base_name": row["基本名"], "short_name": row["略称"], "school_no": to_safe_int(row["No"]), "password": row["Password"], "principal": row["校長名"]})
            save_auth(auth); st.success("保存完了")
        st.divider()
        with st.expander("🗑️ 学校アカウントの削除", expanded=False):
            del_opts = {f"{v['base_name']} ({k})": k for k, v in auth.items()}; t_name = st.selectbox("削除する学校を選択", list(del_opts.keys()))
            if st.button("完全削除する", type="primary") and st.checkbox("理解して削除します"):
                t_sid = del_opts[t_name]; create_backup(); master = load_members_master(force_reload=True); save_members_master(master[master['school_id'] != t_sid])
                if t_sid in auth: del auth[t_sid]; save_auth(auth)
                st.success("削除しました"); time.sleep(1); st.rerun()

    elif admin_tab == "📅 年次処理":
        st.subheader("🌸 年度更新処理")
        if st.button("新年度を開始する"): st.success(perform_year_rollover())
        st.subheader("🎓 卒業生データ")
        grad_df = get_graduates_df()
        if not grad_df.empty:
            out = io.BytesIO(); pd.DataFrame(grad_df).to_excel(out, index=False); st.download_button("ダウンロード", out.getvalue(), "graduates.xlsx")
            if st.button("🗑️ 全て削除"): clear_graduates_archive(); st.success("削除完了")
        else: st.caption("なし")
        st.subheader("⏪ 復元")
        if st.button("バックアップから復元"): st.warning(restore_from_backup())

def main():
    st.set_page_config(page_title="Entry System", layout="wide")
    st.title("🥋 高体連空手エントリーシステム")
    nav = st.radio("Nav", ["🏠 学校ログイン", "🆕 新規登録", "🔧 管理者"], horizontal=True, label_visibility="collapsed")
    auth = load_auth()
    if nav == "🏠 学校ログイン":
        if "logged_in_school" in st.session_state: school_page(st.session_state["logged_in_school"])
        else:
            with st.form("login_form"):
                sorted_auth = sorted(auth.items(), key=lambda x: to_safe_int(x[1].get('school_no', 999)))
                name_map = {f"{v.get('base_name')}高等学校": k for k, v in sorted_auth}
                ph = "（こちらから選択。ない場合は新規登録をしてください）"
                s_name = st.selectbox("学校名", [ph] + list(name_map.keys()))
                pw = st.text_input("パスワード", type="password")
                if st.form_submit_button("ログイン"):
                    if s_name != ph and name_map.get(s_name) and auth[name_map[s_name]]["password"] == pw:
                        st.session_state["logged_in_school"] = name_map[s_name]; st.rerun()
                    elif s_name == ph: st.error("❌ 学校を選択してください")
                    else: st.error("❌ パスワードが違います")
    elif nav == "🆕 新規登録":
        with st.form("reg"):
            bn = st.text_input("学校名 (「高等学校」不要)"); p = st.text_input("校長名"); pw = st.text_input("PW", type="password")
            if st.form_submit_button("登録"):
                if bn and pw:
                    nid = generate_school_id(); auth[nid] = {"base_name": bn, "password": pw, "principal": p, "school_no": 999, "advisors": []}
                    save_auth(auth); st.success("完了"); st.rerun()
    elif nav == "🔧 管理者": admin_page()

if __name__ == "__main__": main()
